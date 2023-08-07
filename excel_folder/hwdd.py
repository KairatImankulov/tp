from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Series, Reference


# e = Workbook()

# e.save("excel_file.xlsx")

# excel = load_workbook("excel_file.xlsx")

wb = Workbook()
ws = wb.active
treeData = [
    ["Type", "Leaf Color", "Height"], 
    ["Maple", "Red", 549], 
    ["Oak", "Green", 783], 
    ["Pine", "Green", 1204]
]
for row in treeData:
    ws.append(row)

ft = Font(bold=True)
for row in ws["A1:C1"]:
    for cell in row:
        cell.font = ft

chart = BarChart()
chart.type = "col"
chart.title = "munar Height"
chart.y_axis.title = 'Height (cm)'
chart.x_axis.title = 'Tree Type'
chart.legend = None

data = Reference(ws, min_col=3, min_row=2, max_row=4, max_col=3)
categories = Reference(ws, min_col=1, min_row=2, max_row=4, max_col=1)

chart.add_data(data)
chart.set_categories(categories)

ws.add_chart(chart, "E5")
wb.save("TreeData.xlsx")

# page = excel['Sheet']

# page["A5"] = 'Среднее'

# s = page['B2'].value + page['B3'].value + page['B4'].value

# medium = s / 3

# page['B5'] = medium

# excel.save("excel_file.xlsx")